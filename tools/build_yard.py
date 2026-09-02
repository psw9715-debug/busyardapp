# -*- coding: utf-8 -*-
"""차고지 프린트.xlsx -> src/yard-data.js

두 시트를 합쳐 배치도 하나를 만든다.
- `신차고지-순서` : 자리마다 순회 번호가 적힌 시트. 앱의 입력 순서 기준.
                    자리 수는 이 시트가 정한다 (번호가 1부터 연속이기만 하면 된다).
- `신차고지`      : 실제로 손으로 적는 인쇄용 시트. 주차장 바닥에 칠해진
                    고정번호가 L열(위 7칸)·M열(아래 13칸)에 따로 적혀 있다.

자리 한 칸 = 엑셀 세로 3행 블록.
  그리드 행 = (엑셀행 - 3) / 3 + 1   (1..20)
  그리드 열 = 엑셀 열 번호            (1..15, A..O)
"""
import json
import openpyxl
from openpyxl.utils import get_column_letter as col_letter

from xlsx_color import load_theme, cell_bg, border_weights

XLSX = "reference/차고지 프린트.xlsx"
OUT = "src/yard-data.js"

ORDER_SHEET = "신차고지-순서"
PRINT_SHEET = "신차고지"

COL_L, COL_M = 12, 13


def grid_row(excel_row):
    return (excel_row - 3) // 3 + 1


def join_label(text):
    """원본에서 줄바꿈된 라벨을 한 줄로 합친다.

    "수\\n소\\n차\\n고\\n지" 는 세로로 쓴 한 단어이므로 그냥 붙이고,
    "B2\\nB4" 는 별개의 두 표지라 구분자를 넣는다.
    """
    lines = [ln for ln in text.split("\n") if ln.strip()]
    if len(lines) <= 1:
        return text.replace("\n", "")
    sep = "/" if all(ln.strip().isalnum() and ln.isascii() for ln in lines) else ""
    return sep.join(ln.strip() for ln in lines)


def main():
    wb = openpyxl.load_workbook(XLSX)
    ws = wb[ORDER_SHEET]
    wp = wb[PRINT_SHEET]

    theme = load_theme(XLSX)

    cells = []
    for rng in ws.merged_cells.ranges:
        if rng.min_row < 3:
            continue  # 1~2행은 제목
        value = ws.cell(rng.min_row, rng.min_col).value
        cell = {
            "col": rng.min_col,
            "colspan": rng.max_col - rng.min_col + 1,
            "row": grid_row(rng.min_row),
            "rowspan": (rng.max_row - rng.min_row + 1) // 3,
            "xl": col_letter(rng.min_col) + str(rng.min_row),
        }
        # 색과 테두리는 실제로 손에 쥐는 인쇄용 시트에서 가져온다
        bg = cell_bg(wp.cell(rng.min_row, rng.min_col), theme)
        if bg:
            cell["bg"] = bg
        cell["b"] = border_weights(wp, rng.min_row, rng.min_col, rng.max_row, rng.max_col)

        if isinstance(value, int):
            cell["kind"] = "spot"
            cell["spot"] = value
        elif value is not None:
            cell["kind"] = "label"
            cell["text"] = join_label(str(value))
        elif cell["col"] == COL_L and cell["row"] >= 8:
            # L열 아래쪽 13칸: 바닥 고정번호만 있고 순회 대상은 아니다.
            # 번호 자체는 옆 M열에 따로 찍히므로 여기는 빈 칸으로 둔다.
            cell["kind"] = "skip"
        else:
            cell["kind"] = "void"
        cells.append(cell)

    # 병합에 속하지 않은 나머지 칸. 여기에 주차장 바닥 고정번호가 있고,
    # 색만 칠해진 구역 표시도 있다 (교통사업처 아래 초록 블록 등).
    # 하나라도 빠지면 인쇄물이 원본과 달라지므로 격자 전체를 훑는다.
    taken = set()
    for c in cells:
        for dr in range(c["rowspan"]):
            for dc in range(c["colspan"]):
                taken.add((c["row"] + dr, c["col"] + dc))

    for excel_row in range(3, 63, 3):
        row = grid_row(excel_row)
        for col in range(1, 16):
            if (row, col) in taken:
                continue
            v = wp.cell(excel_row, col).value
            bg = cell_bg(wp.cell(excel_row, col), theme)
            borders = border_weights(wp, excel_row, col, excel_row + 2, col)
            if v is None and not bg and not any(borders):
                continue  # 정말 아무것도 없는 여백
            c = {
                "col": col,
                "colspan": 1,
                "row": row,
                "rowspan": 1,
                "xl": col_letter(col) + str(excel_row),
                "kind": "paint" if v is not None else "plain",
                "b": borders,
            }
            if v is not None:
                c["text"] = str(v)
            if bg:
                c["bg"] = bg
            cells.append(c)

    cells.sort(key=lambda c: (c["row"], c["col"]))
    spots = sorted(c["spot"] for c in cells if c["kind"] == "spot")

    # 자리 수는 엑셀이 정한다. 나중에 구역이 더 늘어도 이 파일은 그대로 둔다.
    assert spots, "순회 번호가 하나도 없다"
    assert spots == list(range(1, len(spots) + 1)), (
        f"순회 번호가 1부터 연속이 아님 — 빠지거나 겹친 번호가 있다: "
        f"{spots[:5]}...{spots[-5:]}")

    # 같은 칸을 두 번 그리지 않는지 확인
    seen = set()
    for c in cells:
        key = (c["col"], c["row"])
        assert key not in seen, f"{c['xl']} 자리가 겹침"
        seen.add(key)

    data = {
        "id": "new",
        "name": str(wp["A1"].value or "").strip() or "6차고지 (신차고지)",
        "cols": 15,
        "rows": 20,
        "wideCol": 1,        # A열은 원본에서 폭이 2배 이상
        "totalSpots": len(spots),
        "cells": cells,
    }

    body = json.dumps(data, ensure_ascii=False, separators=(",", ":"))
    with open(OUT, "w", encoding="utf-8") as f:
        f.write("// 자동 생성 파일 — 직접 수정하지 말 것. `python tools/build_yard.py` 로 재생성.\n")
        f.write("export const YARD = " + body + ";\n")

    kinds = {}
    for c in cells:
        kinds[c["kind"]] = kinds.get(c["kind"], 0) + 1
    print(f"{OUT} 생성 완료 — {kinds}")


if __name__ == "__main__":
    main()
